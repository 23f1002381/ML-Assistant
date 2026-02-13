"""Excel export module for structured business card data."""

import os
from typing import List, Dict, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from src.utils import logger, OUTPUT_DIR, ENTITY_FIELDS, DEFAULT_EXCEL_FILENAME


def export_to_excel(
    data: List[Dict[str, str]],
    filename: Optional[str] = None,
) -> Optional[str]:
    """Export extracted business card data to an Excel file.

    Args:
        data: List of dictionaries, each containing entity field values.
        filename: Optional custom filename for the Excel file.

    Returns:
        Optional[str]: Path to the created Excel file, or None on failure.
    """
    if not data:
        logger.warning("No data provided for Excel export")
        return None

    if filename is None:
        filename = DEFAULT_EXCEL_FILENAME

    filepath = os.path.join(OUTPUT_DIR, filename)

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Business Cards"

        header_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="2F5496", end_color="2F5496", fill_type="solid"
        )
        header_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell_font = Font(name="Calibri", size=11)
        cell_alignment = Alignment(vertical="top", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        headers = ["#"] + ENTITY_FIELDS
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        for row_idx, record in enumerate(data, 2):
            ws.cell(row=row_idx, column=1, value=row_idx - 1).border = thin_border
            ws.cell(row=row_idx, column=1).font = cell_font
            ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center")

            for col_idx, field in enumerate(ENTITY_FIELDS, 2):
                cell = ws.cell(
                    row=row_idx, column=col_idx, value=record.get(field, "")
                )
                cell.font = cell_font
                cell.alignment = cell_alignment
                cell.border = thin_border

        column_widths = {"#": 5}
        for field in ENTITY_FIELDS:
            if field in ("Address", "Company"):
                column_widths[field] = 35
            elif field in ("Email", "Website"):
                column_widths[field] = 30
            elif field == "Phone":
                column_widths[field] = 25
            else:
                column_widths[field] = 20

        for col_idx, header in enumerate(headers, 1):
            ws.column_dimensions[
                ws.cell(row=1, column=col_idx).column_letter
            ].width = column_widths.get(header, 20)

        ws.auto_filter.ref = ws.dimensions
        wb.save(filepath)
        logger.info("Excel file exported: %s (%d records)", filepath, len(data))
        return filepath

    except Exception as e:
        logger.error("Excel export failed: %s", str(e))
        return None


def export_to_excel_bytes(data: List[Dict[str, str]]) -> Optional[bytes]:
    """Export data to Excel and return as bytes for download.

    Args:
        data: List of dictionaries, each containing entity field values.

    Returns:
        Optional[bytes]: Excel file content as bytes, or None on failure.
    """
    if not data:
        logger.warning("No data provided for Excel export")
        return None

    filepath = export_to_excel(data)
    if filepath is None:
        return None

    try:
        with open(filepath, "rb") as f:
            return f.read()
    except Exception as e:
        logger.error("Failed to read Excel file: %s", str(e))
        return None
